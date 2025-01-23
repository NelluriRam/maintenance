from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from datetime import datetime
import openpyxl
from reportlab.lib.pagesizes import landscape, letter
from reportlab.pdfgen import canvas
import os
from pathlib import Path
from typing import List, Optional

# Create FastAPI app
app = FastAPI(title="Maintenance Work Order System")

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create directories if they don't exist
UPLOAD_DIR = Path("uploads")
REPORTS_DIR = Path("reports")
STATIC_DIR = Path("static")

for dir_path in [UPLOAD_DIR, REPORTS_DIR, STATIC_DIR]:
    dir_path.mkdir(exist_ok=True)

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")


# Pydantic models
class WorkOrder(BaseModel):
    property_code: str
    room_numbers: str
    work_orders: str
    completion_date: str


class RemoveWorkOrder(BaseModel):
    property_code: str
    room_number: str


# Utility functions
def ensure_excel_file(property_code: str) -> Path:
    """Create or return path to Excel file for property."""
    file_path = UPLOAD_DIR / f"{property_code}_work_orders.xlsx"
    if not file_path.exists():
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Work Orders"
        # Updated headers to include Best Room column
        sheet.append(["Room Number", "Work Order", "Completion Date", "Status", "Best Room"])
        wb.save(file_path)
    return file_path


def get_property_name(property_code: str) -> str:
    """Return property name based on code."""
    return {
        "NY198": "Comfort Inn & Suites",
        "NY345": "Quality Inn & Suites"
    }.get(property_code, "Unknown Property")


def format_cell_date(cell) -> str:
    """Format date cell value to string."""
    if isinstance(cell, datetime):
        return cell.strftime("%Y-%m-%d")
    return str(cell)


@app.get("/")
async def read_root():
    """Serve the main HTML page."""
    return FileResponse("static/index.html")


@app.post("/api/work-orders")
async def create_work_order(work_order: WorkOrder):
    """Create new work orders."""
    try:
        # Validate inputs
        room_numbers_list = [room.strip() for room in work_order.room_numbers.split(",")]
        work_orders_list = [work.strip() for work in work_order.work_orders.split(",")]

        if len(room_numbers_list) != len(work_orders_list):
            raise HTTPException(
                status_code=400,
                detail="Number of rooms and work orders must match"
            )

        # Parse completion date
        try:
            completion_date = datetime.strptime(work_order.completion_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="Invalid date format. Use YYYY-MM-DD"
            )

        # Get or create Excel file
        file_path = ensure_excel_file(work_order.property_code)
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Process each work order
        for room_number, work_order_text in zip(room_numbers_list, work_orders_list):
            # Check for existing work order
            existing_row = None
            for row in sheet.iter_rows(min_row=2):
                if str(row[0].value) == room_number:
                    existing_row = row
                    break

            if existing_row:
                # Update existing work order
                current_work_order = str(existing_row[1].value)
                new_work_order = f"{current_work_order} / {work_order_text}"
                existing_row[1].value = new_work_order
                existing_row[2].value = completion_date
            else:
                # Create new work order
                sheet.append([
                    room_number,
                    work_order_text,
                    completion_date,
                    "Pending"
                ])

        # Save changes
        wb.save(file_path)
        return {"status": "success", "message": "Work orders saved successfully"}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/remove-work-order")
async def remove_work_order(remove_order: RemoveWorkOrder):
    """Remove a work order for a specific room."""
    try:
        file_path = UPLOAD_DIR / f"{remove_order.property_code}_work_orders.xlsx"
        if not file_path.exists():
            raise HTTPException(
                status_code=404,
                detail="No work orders found for this property"
            )

        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Find and remove the work order
        row_to_delete = None
        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value) == remove_order.room_number:
                row_to_delete = row[0].row
                break

        if row_to_delete:
            sheet.delete_rows(row_to_delete)
            wb.save(file_path)
            return {
                "status": "success",
                "message": f"Work order for room {remove_order.room_number} removed successfully"
            }
        else:
            raise HTTPException(
                status_code=404,
                detail=f"No work order found for room {remove_order.room_number}"
            )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/generate-report/{property_code}")
async def generate_report(property_code: str):
    """Generate PDF report for property work orders and display it inline."""
    try:
        # Check for work orders file
        file_path = UPLOAD_DIR / f"{property_code}_work_orders.xlsx"
        if not file_path.exists():
            raise HTTPException(
                status_code=404,
                detail="No work orders found for this property"
            )

        # Prepare PDF file
        pdf_path = REPORTS_DIR / f"{property_code}_maintenance_report.pdf"
        property_name = get_property_name(property_code)

        # Create PDF
        c = canvas.Canvas(str(pdf_path), pagesize=landscape(letter))
        width, height = landscape(letter)

        # Add header
        c.setFont("Helvetica-Bold", 18)
        c.drawString(30, height - 50, f"Maintenance Report for {property_name}")
        c.setFont("Helvetica", 12)
        c.drawString(30, height - 70, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        # Draw header line
        c.line(30, height - 85, width - 30, height - 85)

        # Set up table headers
        c.setFont("Helvetica-Bold", 12)
        y_position = height - 120
        col_positions = [30, 130, 380, 530]
        headers = ["Room", "Work Order", "Completion Date", "Status"]

        for pos, header in zip(col_positions, headers):
            c.drawString(pos, y_position, header)

        # Add line under headers
        c.line(30, y_position - 15, width - 30, y_position - 15)

        # Load work orders
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Add work orders to PDF
        y_position -= 40
        c.setFont("Helvetica", 10)

        for row in sheet.iter_rows(min_row=2):
            if y_position < 50:  # New page if needed
                c.showPage()
                c.setFont("Helvetica", 10)
                y_position = height - 50

            # Write row data
            c.drawString(col_positions[0], y_position, str(row[0].value))

            # Wrap work order text
            work_order_text = str(row[1].value)
            words = work_order_text.split()
            lines = []
            current_line = []

            for word in words:
                current_line.append(word)
                if c.stringWidth(' '.join(current_line), "Helvetica", 10) > 230:
                    current_line.pop()
                    lines.append(' '.join(current_line))
                    current_line = [word]

            if current_line:
                lines.append(' '.join(current_line))

            # Draw wrapped text
            for i, line in enumerate(lines):
                c.drawString(col_positions[1], y_position - (i * 12), line)

            # Continue with other columns
            c.drawString(col_positions[2], y_position, format_cell_date(row[2].value))
            c.drawString(col_positions[3], y_position, str(row[3].value))

            y_position -= max(len(lines) * 12, 20)

        # Save PDF
        c.save()

        # Return PDF file with Content-Disposition: inline for browser display
        return FileResponse(
            pdf_path,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"inline; filename={property_code}_maintenance_report.pdf"
            }
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# Project structure helper
def create_project_structure():
    """Create necessary directories and files for the project."""
    # Create directories
    for dir_path in [UPLOAD_DIR, REPORTS_DIR, STATIC_DIR]:
        dir_path.mkdir(exist_ok=True)


@app.get("/api/work-orders/{property_code}")
async def get_work_orders(property_code: str):
    """Get all work orders for a property."""
    try:
        file_path = UPLOAD_DIR / f"{property_code}_work_orders.xlsx"
        if not file_path.exists():
            return {"work_orders": []}

        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        work_orders = []
        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is not None for cell in row[:4]):  # Ensure row has data
                work_orders.append({
                    "room_number": str(row[0].value),
                    "work_order": str(row[1].value),
                    "completion_date": format_cell_date(row[2].value),
                    "status": str(row[3].value),
                    "best_room": str(row[4].value) if len(row) > 4 and row[4].value else "No"
                })

        return {"work_orders": work_orders}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
@app.post("/api/edit-work-order")
async def edit_work_order(request: Request):
    """Edit an existing work order."""
    try:
        data = await request.json()
        property_code = data.get('property_code')
        room_number = data.get('room_number')
        work_order = data.get('work_order')
        completion_date = data.get('completion_date')

        if not all([property_code, room_number, work_order, completion_date]):
            raise HTTPException(status_code=400, detail="Missing required fields")

        try:
            completion_date = datetime.strptime(completion_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="Invalid date format. Use YYYY-MM-DD"
            )

        file_path = UPLOAD_DIR / f"{property_code}_work_orders.xlsx"
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Property not found")

        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Find and update the work order
        found = False
        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value) == room_number:
                row[1].value = work_order  # Update work order
                row[2].value = completion_date  # Update completion date
                found = True
                break

        if not found:
            raise HTTPException(status_code=404, detail="Room not found")

        wb.save(file_path)
        return {"status": "success", "message": "Work order updated successfully"}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/update-room-status")
async def update_room_status(request: Request):
    """Update room status and best room designation."""
    try:
        # Get JSON data from request body
        data = await request.json()
        property_code = data.get('property_code')
        room_number = data.get('room_number')
        status = data.get('status')
        best_room = data.get('best_room', 'No')

        if not all([property_code, room_number, status]):
            raise HTTPException(status_code=400, detail="Missing required fields")

        file_path = UPLOAD_DIR / f"{property_code}_work_orders.xlsx"
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Property not found")

        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Find and update the room
        found = False
        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value) == room_number:
                row[3].value = status  # Update status
                if len(row) > 4:
                    row[4].value = best_room
                else:
                    # Add best_room column if it doesn't exist
                    sheet.cell(row=row[0].row, column=5, value=best_room)
                found = True
                break

        if not found:
            raise HTTPException(status_code=404, detail="Room not found")

        wb.save(file_path)
        return {"status": "success", "message": "Room status updated successfully"}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn

    create_project_structure()
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)